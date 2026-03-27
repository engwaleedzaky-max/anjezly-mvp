from __future__ import annotations

from datetime import datetime
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import PROVIDERS_XLSX, REQUESTS_XLSX
from models import ChatState


def _ensure_header(ws: Worksheet, headers: Sequence[str]) -> None:
    first_row = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
    if all(v is None for v in first_row):
        for i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=header)


def save_request_to_excel(state: ChatState) -> None:
    created_at = datetime.utcnow().isoformat(timespec="seconds")

    if REQUESTS_XLSX.exists():
        wb = load_workbook(REQUESTS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "requests"

    _ensure_header(
        ws,
        ["created_at", "القسم", "الخدمة", "الاسم", "الهاتف", "العنوان", "التفاصيل", "source"],
    )
    ws.append(
        [
            created_at,
            state.category_name,
            state.service_name,
            state.name,
            state.phone,
            state.address,
            state.details,
            "web_chat",
        ]
    )
    wb.save(REQUESTS_XLSX)


def save_provider_to_excel(state: ChatState) -> None:
    created_at = datetime.utcnow().isoformat(timespec="seconds")

    if PROVIDERS_XLSX.exists():
        wb = load_workbook(PROVIDERS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "providers"

    _ensure_header(
        ws,
        ["created_at", "الاسم", "الهاتف", "المهنة", "ماذا تضيف للفريق", "تصنع ايه من البيت", "source"],
    )
    ws.append(
        [
            created_at,
            state.p_name,
            state.p_phone,
            state.p_profession,
            state.p_contrib,
            state.p_home,
            "web_chat",
        ]
    )
    wb.save(PROVIDERS_XLSX)
