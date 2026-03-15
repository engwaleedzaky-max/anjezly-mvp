# file: app.py
"""
Anjezly - Web Chatbot MVP (Bilingual UI + Full Shield Logo + Start over to Language + Back)
- Landing message bilingual (Arabic + English): brand + slogan + language prompt
- Buttons bilingual: Admin Panel, Send, Back, Start over
- Start over always returns to language selection
- Admin PIN = 4321 (or env ADMIN_PIN)
- Customer flow saves to requests.xlsx
- Provider flow saves to providers.xlsx
- Back: one step back
"""

from __future__ import annotations

import os
import re
import secrets
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Literal, Optional

from fastapi import FastAPI, Form, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from starlette.middleware.sessions import SessionMiddleware

# -----------------------------
# Config
# -----------------------------
BRAND_AR = "أنجزلي"
BRAND_EN = "Anjezly"
SLOGAN_AR = "اطلبها، وإحنا ننجزها"
SLOGAN_EN = "Request it, and we’ll get it done."

REQUESTS_XLSX = Path("requests.xlsx")
PROVIDERS_XLSX = Path("providers.xlsx")

# ✅ Admin PIN
ADMIN_PIN = os.environ.get("ADMIN_PIN", "4321")
SESSION_SECRET = os.environ.get("SESSION_SECRET", secrets.token_urlsafe(32))

Lang = Literal["ar", "en"]
Role = Literal["customer", "provider"]

# UI internal commands
CMD_BACK = "__back__"
CMD_RESTART = "__restart__"

# -----------------------------
# Menus (Arabic + English)
# -----------------------------
MAIN_MENU_AR: dict[str, str] = {
    "1": "العقارات والتمويل العقاري",
    "2": "التشطيبات والديكورات",
    "3": "الأجهزة المنزلية",
    "4": "التنقلات والرحلات",
    "5": "اركب مع زميلك",
}
SUB_MENU_AR: dict[str, dict[str, str]] = {
    "1": {"1": "تمليك", "2": "إيجار", "3": "شاليهات", "4": "تسويق عقاري", "5": "تمويل عقاري"},
    "2": {
        "1": "ديكورات وتشطيبات",
        "2": "سمارت هوم ومودرن هوم",
        "3": "تأسيس وتشطيب",
        "4": "سباكة",
        "5": "كهرباء",
        "6": "محارة",
        "7": "جبس بورد وبديل",
        "8": "نقاشة",
        "9": "رخام",
    },
    "3": {"1": "تكييفات", "2": "ثلاجات", "3": "غسالات", "4": "شاشات", "5": "أجهزة مطبخ"},
    "4": {
        "1": "سيارات ملاكي للتنقل",
        "2": "سيارات عائلية للرحلات",
        "3": "حجز شاليهات",
        "4": "رحلات الأقصر",
        "5": "رحلات أسوان",
        "6": "رحلات الغردقة",
        "7": "رحلات شرم الشيخ",
        "8": "رحلات القاهرة",
        "9": "أفريكانو بارك",
    },
    "5": {
        "1": "المعمورة البلد",
        "2": "المندرة",
        "3": "سيدي بشر",
        "4": "رشدي",
        "5": "سموحة",
        "6": "عوايد",
        "7": "كوبري أبيس سريع",
        "8": "البيضة",
        "9": "كفر الدوار",
    },
}

MAIN_MENU_EN: dict[str, str] = {
    "1": "Real Estate & Mortgage",
    "2": "Finishing & Decor",
    "3": "Home Appliances",
    "4": "Transport & Trips",
    "5": "Carpool with Colleagues",
}
SUB_MENU_EN: dict[str, dict[str, str]] = {
    "1": {"1": "Buy", "2": "Rent", "3": "Chalets", "4": "Real Estate Marketing", "5": "Mortgage"},
    "2": {
        "1": "Decor & Finishing",
        "2": "Smart / Modern Home",
        "3": "Setup & Finishing",
        "4": "Plumbing",
        "5": "Electrical",
        "6": "Plastering",
        "7": "Gypsum / Alternatives",
        "8": "Painting",
        "9": "Marble",
    },
    "3": {"1": "Air Conditioners", "2": "Refrigerators", "3": "Washing Machines", "4": "TVs", "5": "Kitchen Appliances"},
    "4": {
        "1": "Sedan Transport",
        "2": "Family Trips",
        "3": "Chalet Booking",
        "4": "Luxor Trips",
        "5": "Aswan Trips",
        "6": "Hurghada Trips",
        "7": "Sharm El Sheikh Trips",
        "8": "Cairo Trips",
        "9": "Africano Park",
    },
    "5": {
        "1": "Al Maamoura",
        "2": "Al Mandara",
        "3": "Sidi Bishr",
        "4": "Roushdy",
        "5": "Smouha",
        "6": "Awayed",
        "7": "Abis Bridge (Fast)",
        "8": "Al Beida",
        "9": "Kafr El Dawar",
    },
}

TXT = {
    "ar": {
        "choose_role": "اختر نوع المستخدم:\n1) طالب خدمة\n2) مقدم خدمة\n",
        "use_buttons": "استخدم زر (رجوع) أو (Start over).",
        "invalid_choice": "اختيار غير صحيح.",
        "customer_title": "✅ أنت طالب خدمة.",
        "provider_title": "✅ أنت مقدم خدمة.",
        "category_prompt": "اختر القسم بكتابة رقم:",
        "service_prompt": "اختر الخدمة بكتابة رقم:",
        "enter_name": "اكتب اسمك:",
        "enter_phone": "اكتب رقم الهاتف:",
        "enter_address": "اكتب العنوان / المنطقة:",
        "enter_details": "اكتب تفاصيل الطلب:",
        "phone_invalid": "رقم غير صحيح. مثال: 01012345678",
        "too_short_name": "الاسم قصير جدًا. اكتب اسمك مرة أخرى:",
        "too_short_addr": "العنوان قصير جدًا. اكتب العنوان/المنطقة مرة أخرى:",
        "too_short_details": "التفاصيل قصيرة جدًا. اكتب تفاصيل أكثر:",
        "saved_ok": "✅ تم تسجيل طلبك بنجاح!",
        "back_msg": "↩️ رجعنا خطوة للخلف.",
        "no_prev": "لا توجد خطوة سابقة.",
        "provider_intro": "احجز موقعك مع فريق عمل أنجزلي 💪\nسجل بياناتك للانضمام كمقدم خدمة.\n",
        "provider_profession": "اكتب مهنتك / تخصصك (مثال: كهرباء، سباكة...):",
        "provider_contrib": "تقدر تضيف ايه للفريق؟ (خبرة/معدات/سرعة استجابة...):",
        "provider_home": "تعرف تصنع ايه من بيتك؟ (اختياري/اكتب: لا يوجد):",
        "provider_saved": "✅ تم تسجيل بياناتك كمقدم خدمة. سنراجعها ونتواصل معك.",
    },
    "en": {
        "choose_role": "Choose user type:\n1) Service requester\n2) Service provider\n",
        "use_buttons": "Use (Back) or (Start over).",
        "invalid_choice": "Invalid choice.",
        "customer_title": "✅ You are a service requester.",
        "provider_title": "✅ You are a service provider.",
        "category_prompt": "Choose a category (type a number):",
        "service_prompt": "Choose a service (type a number):",
        "enter_name": "Enter your name:",
        "enter_phone": "Enter your phone number:",
        "enter_address": "Enter your area / address:",
        "enter_details": "Enter request details:",
        "phone_invalid": "Invalid phone. Example: 01012345678",
        "too_short_name": "Name is too short. Please enter again:",
        "too_short_addr": "Address is too short. Please enter again:",
        "too_short_details": "Details are too short. Please add more:",
        "saved_ok": "✅ Your request has been submitted!",
        "back_msg": "↩️ Went back one step.",
        "no_prev": "No previous step.",
        "provider_intro": "Reserve your spot with Anjezly team 💪\nRegister to join as a service provider.\n",
        "provider_profession": "Enter your profession (e.g., Electrical, Plumbing...):",
        "provider_contrib": "What can you add to the team? (skills/equipment/experience...):",
        "provider_home": "What can you make from home? (optional / type: none):",
        "provider_saved": "✅ Provider info saved. We will review and contact you.",
    },
}

Step = Literal[
    "lang",
    "role",
    "main_menu",
    "sub_menu",
    "name",
    "phone",
    "address",
    "details",
    "p_name",
    "p_phone",
    "p_profession",
    "p_contrib",
    "p_home",
]


@dataclass
class ChatState:
    lang: Optional[Lang] = None
    role: Optional[Role] = None
    step: Step = "lang"

    # customer
    category_key: str = ""
    category_name: str = ""
    service_key: str = ""
    service_name: str = ""
    name: str = ""
    phone: str = ""
    address: str = ""
    details: str = ""

    # provider
    p_name: str = ""
    p_phone: str = ""
    p_profession: str = ""
    p_contrib: str = ""
    p_home: str = ""

    def to_dict(self) -> dict[str, Any]:
        return self.__dict__.copy()

    @classmethod
    def from_dict(cls, raw: dict[str, Any]) -> "ChatState":
        st = cls()
        for k in st.__dict__.keys():
            if k in raw:
                setattr(st, k, raw[k])

        if st.lang not in (None, "ar", "en"):
            st.lang = None
        if st.role not in (None, "customer", "provider"):
            st.role = None

        valid_steps = {
            "lang",
            "role",
            "main_menu",
            "sub_menu",
            "name",
            "phone",
            "address",
            "details",
            "p_name",
            "p_phone",
            "p_profession",
            "p_contrib",
            "p_home",
        }
        if st.step not in valid_steps:
            st.step = "lang"
        return st


def normalize_digits(s: str) -> str:
    trans = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")
    return (s or "").translate(trans)


def normalize_text(s: str) -> str:
    return (s or "").strip().replace("\u200f", "").replace("\u200e", "")


def validate_phone(raw: str) -> Optional[str]:
    s = normalize_digits(raw)
    s = re.sub(r"\s+", "", s).replace("+", "")
    digits = re.sub(r"\D", "", s)
    return digits if 8 <= len(digits) <= 15 else None


def get_state(request: Request) -> ChatState:
    raw = request.session.get("chat_state")
    return ChatState.from_dict(raw) if isinstance(raw, dict) else ChatState()


def set_state(request: Request, state: ChatState) -> None:
    request.session["chat_state"] = state.to_dict()


def get_history(request: Request) -> list[dict[str, Any]]:
    raw = request.session.get("chat_history")
    return [x for x in raw if isinstance(x, dict)] if isinstance(raw, list) else []


def set_history(request: Request, history: list[dict[str, Any]]) -> None:
    request.session["chat_history"] = history


def clear_chat(request: Request) -> None:
    request.session.pop("chat_state", None)
    request.session.pop("chat_history", None)


def push_history(request: Request, state: ChatState) -> None:
    hist = get_history(request)
    hist.append(state.to_dict())
    set_history(request, hist)


def pop_history(request: Request) -> Optional[ChatState]:
    hist = get_history(request)
    if not hist:
        return None
    last = hist.pop()
    set_history(request, hist)
    return ChatState.from_dict(last)


def menu_for(lang: Lang) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    return (MAIN_MENU_EN, SUB_MENU_EN) if lang == "en" else (MAIN_MENU_AR, SUB_MENU_AR)


def prompt_language_bilingual() -> str:
    return (
        f"{BRAND_AR} | {BRAND_EN} 👋\n"
        f"{SLOGAN_AR}\n{SLOGAN_EN}\n\n"
        "اختر اللغة:\n1) عربي\n2) English\n\n"
        "Choose language:\n1) عربي\n2) English\n\n"
        "Use (Back) or (Start over). / استخدم (رجوع) أو (Start over)."
    )


def prompt_role(lang: Lang) -> str:
    return TXT[lang]["choose_role"] + "\n" + TXT[lang]["use_buttons"]


def prompt_main_menu(lang: Lang) -> str:
    main, _ = menu_for(lang)
    lines = [TXT[lang]["customer_title"], TXT[lang]["category_prompt"], ""]
    for k, v in main.items():
        lines.append(f"{k}) {v}")
    lines.append("")
    lines.append(TXT[lang]["use_buttons"])
    return "\n".join(lines)


def prompt_sub_menu(lang: Lang, category_key: str) -> str:
    main, sub = menu_for(lang)
    cat_name = main.get(category_key, "")
    subs = sub.get(category_key, {})
    lines = [f"✅ {cat_name}", "", TXT[lang]["service_prompt"], ""]
    for k, v in subs.items():
        lines.append(f"{k}) {v}")
    lines.append("")
    lines.append(TXT[lang]["use_buttons"])
    return "\n".join(lines)


def excel_ensure_header(ws: Worksheet, headers: list[str]) -> None:
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(headers)


def save_request_to_excel(state: ChatState) -> None:
    created_at = datetime.utcnow().isoformat(timespec="seconds")
    if REQUESTS_XLSX.exists():
        wb = load_workbook(REQUESTS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "requests"

    excel_ensure_header(
        ws,
        ["created_at", "lang", "category", "service", "name", "phone", "address", "details", "source"],
    )
    ws.append(
        [
            created_at,
            state.lang,
            state.category_name,
            state.service_name,
            state.name,
            state.phone,
            state.address,
            state.details,
            "web_chat",
        ]
    )
    wb.save(REQUESTS_XLSX)


def save_provider_to_excel(state: ChatState) -> None:
    created_at = datetime.utcnow().isoformat(timespec="seconds")
    if PROVIDERS_XLSX.exists():
        wb = load_workbook(PROVIDERS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "providers"

    excel_ensure_header(
        ws,
        ["created_at", "lang", "name", "phone", "profession", "team_contribution", "home_made", "source"],
    )
    ws.append(
        [
            created_at,
            state.lang,
            state.p_name,
            state.p_phone,
            state.p_profession,
            state.p_contrib,
            state.p_home,
            "web_chat",
        ]
    )
    wb.save(PROVIDERS_XLSX)


def choose_lang(text: str) -> Optional[Lang]:
    t = normalize_digits(text).strip().lower()
    if t in {"1", "ar", "arabic", "عربي"}:
        return "ar"
    if t in {"2", "en", "english"}:
        return "en"
    return None


def choose_role(text: str) -> Optional[Role]:
    t = normalize_digits(text).strip().lower()
    if t in {"1", "customer", "requester"}:
        return "customer"
    if t in {"2", "provider"}:
        return "provider"
    return None


def choose_category(lang: Lang, text: str) -> Optional[str]:
    main, _ = menu_for(lang)
    t = normalize_digits(text).strip()
    return t if t in main else None


def choose_service(lang: Lang, category_key: str, text: str) -> Optional[str]:
    _, sub = menu_for(lang)
    t = normalize_digits(text).strip()
    return t if t in sub.get(category_key, {}) else None


def handle_restart_to_language(request: Request) -> str:
    clear_chat(request)
    st = ChatState(lang=None, role=None, step="lang")
    set_state(request, st)
    set_history(request, [])
    return "🔄 Start over.\n\n" + prompt_language_bilingual()


def prompt_for_step(state: ChatState) -> str:
    if state.step == "lang" or state.lang is None:
        return prompt_language_bilingual()

    lang = state.lang
    if state.step == "role":
        return prompt_role(lang)

    if state.step == "main_menu":
        return prompt_main_menu(lang)
    if state.step == "sub_menu":
        return prompt_sub_menu(lang, state.category_key)
    if state.step == "name":
        return f"✅ {state.service_name}\n\n{TXT[lang]['enter_name']}"
    if state.step == "phone":
        return TXT[lang]["enter_phone"]
    if state.step == "address":
        return TXT[lang]["enter_address"]
    if state.step == "details":
        return f"{TXT[lang]['enter_details']} ({state.service_name})"

    if state.step == "p_name":
        return TXT[lang]["provider_intro"] + "\n" + TXT[lang]["enter_name"]
    if state.step == "p_phone":
        return TXT[lang]["enter_phone"]
    if state.step == "p_profession":
        return TXT[lang]["provider_profession"]
    if state.step == "p_contrib":
        return TXT[lang]["provider_contrib"]
    if state.step == "p_home":
        return TXT[lang]["provider_home"]

    return prompt_language_bilingual()


def handle_back(request: Request) -> str:
    prev = pop_history(request)
    cur = get_state(request)
    if not prev:
        return "No previous step. / لا توجد خطوة سابقة.\n\n" + prompt_for_step(cur)

    set_state(request, prev)
    if prev.lang:
        return TXT[prev.lang]["back_msg"] + "\n\n" + prompt_for_step(prev)
    return "↩️ Back.\n\n" + prompt_for_step(prev)


def bot_reply(request: Request, user_text: str) -> str:
    text = normalize_text(user_text)
    state = get_state(request)

    if text == CMD_BACK:
        return handle_back(request)
    if text == CMD_RESTART:
        return handle_restart_to_language(request)

    if text.lower() in {"start over", "restart", "#", "ابدأ", "ابدا", "start", "menu", "reset", "إلغاء", "الغاء"}:
        return handle_restart_to_language(request)

    # language step
    if state.step == "lang" or state.lang is None:
        chosen = choose_lang(text)
        if not chosen:
            return "Invalid choice / اختيار غير صحيح.\n\n" + prompt_language_bilingual()
        push_history(request, state)
        state.lang = chosen
        state.step = "role"
        set_state(request, state)
        return prompt_role(chosen)

    lang = state.lang

    # role step
    if state.step == "role":
        chosen = choose_role(text)
        if not chosen:
            return TXT[lang]["invalid_choice"] + "\n\n" + prompt_role(lang)
        push_history(request, state)
        state.role = chosen
        state.step = "main_menu" if chosen == "customer" else "p_name"
        set_state(request, state)
        return prompt_for_step(state)

    # customer flow
    if state.role == "customer":
        if state.step == "main_menu":
            cat_key = choose_category(lang, text)
            if not cat_key:
                return TXT[lang]["invalid_choice"] + "\n\n" + prompt_main_menu(lang)
            push_history(request, state)
            main, _ = menu_for(lang)
            state.category_key = cat_key
            state.category_name = main[cat_key]
            state.step = "sub_menu"
            set_state(request, state)
            return prompt_sub_menu(lang, cat_key)

        if state.step == "sub_menu":
            svc_key = choose_service(lang, state.category_key, text)
            if not svc_key:
                return TXT[lang]["invalid_choice"] + "\n\n" + prompt_sub_menu(lang, state.category_key)
            push_history(request, state)
            _, sub = menu_for(lang)
            state.service_key = svc_key
            state.service_name = sub[state.category_key][svc_key]
            state.step = "name"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "name":
            if len(text) < 2:
                return TXT[lang]["too_short_name"]
            push_history(request, state)
            state.name = text
            state.step = "phone"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "phone":
            phone = validate_phone(text)
            if not phone:
                return TXT[lang]["phone_invalid"]
            push_history(request, state)
            state.phone = phone
            state.step = "address"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "address":
            if len(text) < 3:
                return TXT[lang]["too_short_addr"]
            push_history(request, state)
            state.address = text
            state.step = "details"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "details":
            if len(text) < 5:
                return TXT[lang]["too_short_details"]
            state.details = text
            save_request_to_excel(state)

            confirmation = (
                f"{TXT[lang]['saved_ok']}\n\n"
                f"{state.category_name}\n{state.service_name}\n\n"
                f"{TXT[lang]['enter_name']} {state.name}\n"
                f"{TXT[lang]['enter_phone']} {state.phone}\n"
                f"{TXT[lang]['enter_address']} {state.address}\n"
                f"{TXT[lang]['enter_details']} {state.details}\n\n"
            )
            return confirmation + handle_restart_to_language(request)

        return handle_restart_to_language(request)

    # provider flow
    if state.role == "provider":
        if state.step == "p_name":
            if len(text) < 2:
                return TXT[lang]["too_short_name"]
            push_history(request, state)
            state.p_name = text
            state.step = "p_phone"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "p_phone":
            phone = validate_phone(text)
            if not phone:
                return TXT[lang]["phone_invalid"]
            push_history(request, state)
            state.p_phone = phone
            state.step = "p_profession"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "p_profession":
            if len(text) < 2:
                return TXT[lang]["invalid_choice"] + "\n\n" + TXT[lang]["provider_profession"]
            push_history(request, state)
            state.p_profession = text
            state.step = "p_contrib"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "p_contrib":
            if len(text) < 2:
                return TXT[lang]["invalid_choice"] + "\n\n" + TXT[lang]["provider_contrib"]
            push_history(request, state)
            state.p_contrib = text
            state.step = "p_home"
            set_state(request, state)
            return prompt_for_step(state)

        if state.step == "p_home":
            state.p_home = text if text else ("none" if lang == "en" else "لا يوجد")
            save_provider_to_excel(state)

            msg = (
                f"{TXT[lang]['provider_saved']}\n\n"
                f"{TXT[lang]['enter_name']} {state.p_name}\n"
                f"{TXT[lang]['enter_phone']} {state.p_phone}\n"
                f"{TXT[lang]['provider_profession']} {state.p_profession}\n"
                f"{TXT[lang]['provider_contrib']} {state.p_contrib}\n"
                f"{TXT[lang]['provider_home']} {state.p_home}\n\n"
            )
            return msg + handle_restart_to_language(request)

        return handle_restart_to_language(request)

    return handle_restart_to_language(request)


# -----------------------------
# Web UI (no f-string JS issues)
# -----------------------------
HTML_TEMPLATE = r"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>__TITLE__</title>
  <style>
    body { font-family: system-ui, Arial; margin: 0; background:#0b1220; color:#fff; }
    .wrap { max-width: 980px; margin: 0 auto; padding: 18px; }
    .header { display:flex; justify-content: space-between; align-items:center; gap:12px; }
    .brand { display:flex; align-items:center; gap:14px; }
    /* Logo is 3x bigger than old (42 -> 126) */
    .logo { width:126px; height:126px; border-radius:28px; background: rgba(22,163,74,.12);
            display:flex; align-items:center; justify-content:center; overflow:hidden; }
    .logo svg { width:92px; height:92px; }
    .title { margin:0; font-weight:900; font-size: 30px; line-height: 1.1; }
    .slogan { margin:0; opacity:.85; }
    .card { margin-top:14px; background: #0f1a2e; border:1px solid rgba(255,255,255,.08); border-radius: 16px; padding: 14px; }
    .chat { height: 60vh; overflow:auto; padding: 10px; display:flex; flex-direction:column; gap:10px; }
    .bubble { max-width: 88%; padding: 10px 12px; border-radius: 14px; line-height:1.7; white-space: pre-wrap; }
    .bot { background: rgba(255,255,255,.08); align-self:flex-start; }
    .me { background: rgba(22,163,74,.22); align-self:flex-end; }
    .row { display:flex; gap:10px; margin-top: 12px; flex-wrap: wrap; }
    input { flex:1; min-width: 240px; padding: 12px; border-radius: 14px;
            border:1px solid rgba(255,255,255,.12); background:#0b1220; color:#fff; }
    button { padding: 12px 14px; border-radius: 14px; border:none; color:#fff; font-weight:900; cursor:pointer; }
    button:disabled { opacity:.6; cursor:not-allowed; }
    .btn-send { background:#16A34A; }
    .btn-back { background: rgba(255,255,255,.14); }
    .btn-restart { background: rgba(255,255,255,.10); }
    a { color:#a7f3d0; text-decoration:none; }
    .chips { display:flex; gap:8px; flex-wrap: wrap; margin-top: 10px; }
    .chip { padding: 8px 10px; border-radius: 999px; background: rgba(255,255,255,.08);
            border: 1px solid rgba(255,255,255,.10); cursor:pointer; font-weight:800; }
    .chip:hover { background: rgba(255,255,255,.12); }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="header">
      <div class="brand">
        <div class="logo" aria-label="Logo">
          <!-- Full Shield (fills the icon more) -->
          <svg viewBox="0 0 128 128" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M64 6c21 16 40 20 54 23v39c0 34-20 52-54 60C30 120 10 102 10 68V29c14-3 33-7 54-23Z"
                  fill="rgba(22,163,74,.20)" stroke="rgba(226,232,240,.92)" stroke-width="5" stroke-linejoin="round"/>
            <path d="M78 20 36 78h30l-6 34 42-60H72l6-32Z" fill="#16A34A"/>
            <circle cx="96" cy="96" r="20" fill="rgba(15,23,42,.92)" stroke="rgba(226,232,240,.85)" stroke-width="3"/>
            <path d="M87 96l7 8 15-17" stroke="#fff" stroke-width="7" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
        <div>
          <h1 class="title">__BRAND_AR__ | __BRAND_EN__</h1>
          <p class="slogan">__SLOGAN_AR__</p>
          <p class="slogan">__SLOGAN_EN__</p>
        </div>
      </div>
      <div><a href="/admin">لوحة الأدمن / Admin Panel</a></div>
    </div>

    <div class="card">
      <div id="chat" class="chat">
        <div class="bubble bot">__INITIAL__</div>
      </div>

      <div class="row">
        <input id="msg" placeholder="Type a number/message... / اكتب رقم/رسالة..." />
        <button id="send" class="btn-send">Send / إرسال</button>
        <button id="back" class="btn-back" title="Back / رجوع">Back / رجوع</button>
        <button id="restart" class="btn-restart" title="Start over / بدء من جديد">Start over / بدء من جديد</button>
      </div>

      <div class="chips">
        <div class="chip" data-text="1">1</div><div class="chip" data-text="2">2</div><div class="chip" data-text="3">3</div>
        <div class="chip" data-text="4">4</div><div class="chip" data-text="5">5</div><div class="chip" data-text="6">6</div>
        <div class="chip" data-text="7">7</div><div class="chip" data-text="8">8</div><div class="chip" data-text="9">9</div>
      </div>
    </div>
  </div>

<script>
  const CMD_BACK = "__CMD_BACK__";
  const CMD_RESTART = "__CMD_RESTART__";

  const chat = document.getElementById("chat");
  const msg = document.getElementById("msg");
  const send = document.getElementById("send");
  const back = document.getElementById("back");
  const restart = document.getElementById("restart");

  function addBubble(text, who) {
    const div = document.createElement("div");
    div.className = "bubble " + who;
    div.textContent = text;
    chat.appendChild(div);
    chat.scrollTop = chat.scrollHeight;
  }

  async function postText(text) {
    const res = await fetch("/api/message", {
      method: "POST",
      headers: { "Content-Type": "application/x-www-form-urlencoded" },
      body: new URLSearchParams({ text })
    });
    return await res.json();
  }

  async function sendMessage(textOverride=null) {
    const text = (textOverride ?? msg.value).trim();
    if (!text) return;
    msg.value = "";
    addBubble(text, "me");
    send.disabled = true; back.disabled = true; restart.disabled = true;
    try {
      const data = await postText(text);
      addBubble(data.reply, "bot");
    } catch (e) {
      addBubble("Error. / حدث خطأ.", "bot");
    } finally {
      send.disabled = false; back.disabled = false; restart.disabled = false;
      msg.focus();
    }
  }

  async function backStep() {
    send.disabled = true; back.disabled = true; restart.disabled = true;
    try {
      addBubble("Back / رجوع", "me");
      const data = await postText(CMD_BACK);
      addBubble(data.reply, "bot");
    } catch (e) {
      addBubble("Error. / حدث خطأ.", "bot");
    } finally {
      send.disabled = false; back.disabled = false; restart.disabled = false;
      msg.focus();
    }
  }

  async function restartChat() {
    send.disabled = true; back.disabled = true; restart.disabled = true;
    try {
      addBubble("Start over / بدء من جديد", "me");
      const data = await postText(CMD_RESTART);
      addBubble(data.reply, "bot");
    } catch (e) {
      addBubble("Error. / حدث خطأ.", "bot");
    } finally {
      send.disabled = false; back.disabled = false; restart.disabled = false;
      msg.focus();
    }
  }

  send.addEventListener("click", () => sendMessage());
  msg.addEventListener("keydown", (e) => { if (e.key === "Enter") sendMessage(); });
  back.addEventListener("click", backStep);
  restart.addEventListener("click", restartChat);

  document.querySelectorAll(".chip").forEach(el => {
    el.addEventListener("click", () => sendMessage(el.dataset.text));
  });
</script>
</body>
</html>
"""


# -----------------------------
# FastAPI
# -----------------------------
app = FastAPI(title=f"{BRAND_EN} | {BRAND_AR}")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, same_site="lax")


@app.get("/", response_class=HTMLResponse)
def chat_page(request: Request) -> str:
    if not isinstance(request.session.get("chat_state"), dict):
        st = ChatState(lang=None, role=None, step="lang")
        set_state(request, st)
        set_history(request, [])

    state = get_state(request)
    initial = prompt_for_step(state)

    page = (
        HTML_TEMPLATE.replace("__TITLE__", f"{BRAND_EN} | {BRAND_AR}")
        .replace("__BRAND_AR__", BRAND_AR)
        .replace("__BRAND_EN__", BRAND_EN)
        .replace("__SLOGAN_AR__", SLOGAN_AR)
        .replace("__SLOGAN_EN__", SLOGAN_EN)
        .replace("__INITIAL__", initial)
        .replace("__CMD_BACK__", CMD_BACK)
        .replace("__CMD_RESTART__", CMD_RESTART)
    )
    return page


@app.post("/api/message")
def api_message(request: Request, text: str = Form(...)) -> JSONResponse:
    reply = bot_reply(request, text)
    return JSONResponse({"reply": reply})


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request, pin: Optional[str] = None) -> str:
    ok = request.session.get("admin_ok") is True
    if pin and pin.strip() == ADMIN_PIN:
        request.session["admin_ok"] = True
        ok = True

    if not ok:
        return f"""<!doctype html>
<html lang="ar" dir="rtl"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{BRAND_EN} | {BRAND_AR} Admin</title></head>
<body style="font-family:system-ui,Arial; padding:24px;">
<h2>لوحة الأدمن / Admin Panel</h2>
<p>ادخل Admin PIN / Enter Admin PIN</p>
<form method="get" action="/admin">
  <input name="pin" placeholder="PIN" style="padding:10px; width:240px;"/>
  <button type="submit" style="padding:10px;">Login / دخول</button>
</form>
</body></html>"""

    req_exists = REQUESTS_XLSX.exists()
    prov_exists = PROVIDERS_XLSX.exists()

    return f"""<!doctype html>
<html lang="ar" dir="rtl"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{BRAND_EN} | {BRAND_AR} Admin</title></head>
<body style="font-family:system-ui,Arial; padding:24px;">
<h2>لوحة الأدمن / Admin Panel</h2>
<p><a href="/">Back to chat / رجوع للشات</a></p>

<h3>Requests / طلبات طالبي الخدمة</h3>
<p>Status: {"✅ Exists" if req_exists else "❌ Not yet"} / الحالة: {"✅ موجود" if req_exists else "❌ لا يوجد بعد"}</p>
{"<a href='/admin/download/requests'>⬇️ Download requests.xlsx</a>" if req_exists else ""}

<hr style="margin:18px 0;"/>

<h3>Providers / مقدمي الخدمة</h3>
<p>Status: {"✅ Exists" if prov_exists else "❌ Not yet"} / الحالة: {"✅ موجود" if prov_exists else "❌ لا يوجد بعد"}</p>
{"<a href='/admin/download/providers'>⬇️ Download providers.xlsx</a>" if prov_exists else ""}

</body></html>"""


@app.get("/admin/download/{which}")
def admin_download(request: Request, which: str) -> Response:
    if request.session.get("admin_ok") is not True:
        return RedirectResponse("/admin", status_code=303)

    if which == "requests":
        if not REQUESTS_XLSX.exists():
            return RedirectResponse("/admin", status_code=303)
        return FileResponse(str(REQUESTS_XLSX), filename="requests.xlsx")

    if which == "providers":
        if not PROVIDERS_XLSX.exists():
            return RedirectResponse("/admin", status_code=303)
        return FileResponse(str(PROVIDERS_XLSX), filename="providers.xlsx")

    return RedirectResponse("/admin", status_code=303)
